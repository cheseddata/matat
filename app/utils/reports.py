"""Shared report infrastructure.

Every Gemach report needs three output modes:
  1. HTML (Access-style Hebrew table, embedded in the app)
  2. PDF (printable, Hebrew RTL, via ReportLab)
  3. Excel (.xlsx, RTL, via openpyxl)

This module provides the helpers so a new report is ~30 lines:
    from app.utils.reports import export_pdf, export_xlsx, ReportSpec
    spec = ReportSpec(title='...', columns=[...], rows=[...])
    if fmt == 'pdf': return export_pdf(spec)
    if fmt == 'xlsx': return export_xlsx(spec)
    return render_template('gemach/reports/_table.html', spec=spec)
"""
from __future__ import annotations

import io
import os
from dataclasses import dataclass, field
from datetime import datetime, date
from decimal import Decimal
from typing import Any, Callable, Iterable

from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Report spec (everything a report produces)
# ---------------------------------------------------------------------------
@dataclass
class Column:
    """One column in a report table.

    Attributes
    ----------
    key      : dict key / attribute used to pull the value from a row
    title    : Hebrew header label (primary)
    title_en : English header label (fallback)
    width    : relative width for PDF layout (defaults to 1)
    align    : 'right' | 'left' | 'center' — for currency/number, pass 'left' in RTL
    formatter: optional callable(value) -> str, for special formatting
    """
    key: str
    title: str
    title_en: str = ''
    width: float = 1.0
    align: str = 'right'
    formatter: Callable[[Any], str] | None = None


@dataclass
class ReportSpec:
    title: str
    subtitle: str = ''
    columns: list[Column] = field(default_factory=list)
    rows: list[dict] = field(default_factory=list)
    filters: dict = field(default_factory=dict)       # {label: value} shown in header
    totals: dict = field(default_factory=dict)        # {column_key: total_value}
    generated_at: datetime = field(default_factory=datetime.utcnow)
    lang: str = 'he'     # 'he' or 'en'
    rtl: bool = True


# ---------------------------------------------------------------------------
# Formatting helpers
# ---------------------------------------------------------------------------
def fmt_money(v, symbol: str = '') -> str:
    if v is None:
        return ''
    try:
        n = float(v)
    except (TypeError, ValueError):
        return str(v)
    s = f'{n:,.2f}'
    return f'{symbol}{s}' if symbol else s


def fmt_int(v) -> str:
    if v is None:
        return ''
    try:
        return f'{int(v):,}'
    except (TypeError, ValueError):
        return str(v)


def fmt_date(v) -> str:
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%d/%m/%Y')
    if isinstance(v, date):
        return v.strftime('%d/%m/%Y')
    return str(v)


def apply_formatter(column: Column, value: Any) -> str:
    if column.formatter:
        try:
            return column.formatter(value)
        except Exception:
            return str(value) if value is not None else ''
    if value is None:
        return ''
    if isinstance(value, (Decimal, float)):
        return fmt_money(value)
    if isinstance(value, (datetime, date)):
        return fmt_date(value)
    return str(value)


# ---------------------------------------------------------------------------
# Hebrew font (ReportLab) — lazy-registered so import-time is fast
# ---------------------------------------------------------------------------
_HEBREW_FONT_REGISTERED = False
_HEBREW_FONT_NAME = 'HebrewReport'


def _register_hebrew_font():
    """Register a TTF that supports Hebrew glyphs.

    Strategy: try common Windows system fonts in order. On Linux, try DejaVu.
    If nothing is found, ReportLab will fall back to Helvetica which renders
    Hebrew as boxes — ugly but won't crash.
    """
    global _HEBREW_FONT_REGISTERED
    if _HEBREW_FONT_REGISTERED:
        return True

    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    candidates = [
        # Windows
        r'C:\Windows\Fonts\david.ttf',
        r'C:\Windows\Fonts\davidbd.ttf',
        r'C:\Windows\Fonts\arial.ttf',
        r'C:\Windows\Fonts\tahoma.ttf',
        r'C:\Windows\Fonts\segoeui.ttf',
        # Linux / WSL
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
        # macOS
        '/Library/Fonts/Arial.ttf',
        '/System/Library/Fonts/Supplemental/Arial.ttf',
    ]
    for path in candidates:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont(_HEBREW_FONT_NAME, path))
                _HEBREW_FONT_REGISTERED = True
                return True
            except Exception:
                continue
    return False


def _reshape_hebrew(text: str) -> str:
    """Reverse a Hebrew string for display (ReportLab doesn't natively
    handle bidi). This is a crude approach but works for pure-Hebrew
    strings; numbers/English inside Hebrew get ugly.

    For proper bidi, install python-bidi + arabic-reshaper. Marked as TODO
    since adding deps is extra for a sandbox.
    """
    if not text:
        return ''
    # Heuristic: if any Hebrew codepoint, reverse the whole string.
    if any('\u0590' <= c <= '\u05FF' for c in text):
        return text[::-1]
    return text


# ---------------------------------------------------------------------------
# PDF export
# ---------------------------------------------------------------------------
def export_pdf(spec: ReportSpec, filename: str = None):
    """Render the spec to a PDF and return a Flask send_file response."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.units import mm

    _register_hebrew_font()
    font = _HEBREW_FONT_NAME if _HEBREW_FONT_REGISTERED else 'Helvetica'

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        topMargin=15 * mm, bottomMargin=15 * mm,
        leftMargin=12 * mm, rightMargin=12 * mm,
        title=spec.title,
    )

    elements = []

    # Title + meta
    title_style = ParagraphStyle(
        'RTLTitle', fontName=font, fontSize=16, alignment=2 if spec.rtl else 0,
        textColor=colors.HexColor('#2F5597'), spaceAfter=4,
    )
    meta_style = ParagraphStyle(
        'RTLMeta', fontName=font, fontSize=9, alignment=2 if spec.rtl else 0,
        textColor=colors.HexColor('#666666'), spaceAfter=8,
    )

    elements.append(Paragraph(_reshape_hebrew(spec.title), title_style))
    if spec.subtitle:
        elements.append(Paragraph(_reshape_hebrew(spec.subtitle), meta_style))

    meta_bits = [f'הופק: {spec.generated_at:%d/%m/%Y %H:%M}']
    for k, v in (spec.filters or {}).items():
        meta_bits.append(f'{k}: {v}')
    elements.append(Paragraph(_reshape_hebrew(' | '.join(meta_bits)), meta_style))
    elements.append(Spacer(1, 4 * mm))

    # Table
    header = [_reshape_hebrew(c.title or c.title_en) for c in spec.columns]
    # In RTL, reverse the COLUMN order so the "first" column appears on the right.
    if spec.rtl:
        header = list(reversed(header))

    body_rows = []
    for row in spec.rows:
        cells = [apply_formatter(c, row.get(c.key)) for c in spec.columns]
        cells = [_reshape_hebrew(str(x)) for x in cells]
        if spec.rtl:
            cells = list(reversed(cells))
        body_rows.append(cells)

    # Totals row
    if spec.totals:
        totals_cells = []
        for c in spec.columns:
            if c.key in spec.totals:
                totals_cells.append(apply_formatter(c, spec.totals[c.key]))
            else:
                totals_cells.append('')
        # Put "סה״כ" label on the rightmost column (first in RTL)
        if spec.rtl and totals_cells and not totals_cells[0]:
            totals_cells[0] = 'סה״כ'
        totals_cells = [_reshape_hebrew(str(x)) for x in totals_cells]
        if spec.rtl:
            totals_cells = list(reversed(totals_cells))
        body_rows.append(totals_cells)

    col_widths = [c.width for c in spec.columns]
    if spec.rtl:
        col_widths = list(reversed(col_widths))
    total_w = sum(col_widths) or 1.0
    page_w = landscape(A4)[0] - (24 * mm)
    col_widths_abs = [w / total_w * page_w for w in col_widths]

    table = Table([header] + body_rows, colWidths=col_widths_abs, repeatRows=1)
    ts = TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#000080')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2 if spec.totals else -1),
         [colors.white, colors.HexColor('#F2F2F2')]),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
    ])
    if spec.totals:
        ts.add('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D6E4F0'))
        ts.add('FONTSIZE', (0, -1), (-1, -1), 10)
    table.setStyle(ts)
    elements.append(table)

    doc.build(elements)
    buf.seek(0)

    filename = filename or _slug(spec.title) + '.pdf'
    return send_file(buf, mimetype='application/pdf',
                     as_attachment=True, download_name=filename)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------
def export_xlsx(spec: ReportSpec, filename: str = None):
    wb = Workbook()
    ws = wb.active
    ws.title = spec.title[:28] or 'Report'
    if spec.rtl:
        ws.sheet_view.rightToLeft = True

    thin = Side(style='thin', color='BBBBBB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title (merged across columns)
    ncols = len(spec.columns)
    if ncols:
        ws.cell(row=1, column=1, value=spec.title).font = Font(
            name='Segoe UI', size=14, bold=True, color='2F5597')
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        ws.cell(row=1, column=1).alignment = Alignment(
            horizontal='right' if spec.rtl else 'left', vertical='center')

        if spec.subtitle:
            ws.cell(row=2, column=1, value=spec.subtitle).font = Font(
                name='Segoe UI', size=10, color='666666', italic=True)
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)

        meta = f'הופק: {spec.generated_at:%d/%m/%Y %H:%M}'
        for k, v in (spec.filters or {}).items():
            meta += f' | {k}: {v}'
        ws.cell(row=3, column=1, value=meta).font = Font(
            name='Segoe UI', size=9, color='999999')
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)

    # Header row
    header_row = 5
    header_fill = PatternFill('solid', fgColor='000080')
    header_font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
    for i, c in enumerate(spec.columns, start=1):
        cell = ws.cell(row=header_row, column=i, value=c.title or c.title_en)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Body rows
    for r_i, row in enumerate(spec.rows, start=header_row + 1):
        for c_i, c in enumerate(spec.columns, start=1):
            raw = row.get(c.key)
            cell = ws.cell(row=r_i, column=c_i)
            if isinstance(raw, (int, float, Decimal)):
                cell.value = float(raw) if isinstance(raw, Decimal) else raw
                cell.number_format = '#,##0.00'
            elif isinstance(raw, (datetime, date)):
                cell.value = raw
                cell.number_format = 'DD/MM/YYYY'
            else:
                cell.value = '' if raw is None else str(raw)
            cell.alignment = Alignment(
                horizontal='right' if spec.rtl else 'left', vertical='center')
            cell.border = border
            if (r_i - header_row) % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F2F2F2')

    # Totals row
    if spec.totals:
        total_row = header_row + len(spec.rows) + 1
        for c_i, c in enumerate(spec.columns, start=1):
            cell = ws.cell(row=total_row, column=c_i)
            if c.key in spec.totals:
                v = spec.totals[c.key]
                cell.value = float(v) if isinstance(v, Decimal) else v
                cell.number_format = '#,##0.00'
            else:
                # Put "סה״כ" in the first (rightmost in RTL) column
                cell.value = 'סה״כ' if c_i == 1 else ''
            cell.font = Font(name='Segoe UI', bold=True)
            cell.fill = PatternFill('solid', fgColor='D6E4F0')
            cell.alignment = Alignment(
                horizontal='right' if spec.rtl else 'left', vertical='center')
            cell.border = border

    # Column widths
    for i, c in enumerate(spec.columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(40, c.width * 14))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = filename or _slug(spec.title) + '.xlsx'
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


# ---------------------------------------------------------------------------
def _slug(text: str) -> str:
    import re
    s = (text or 'report').strip().replace(' ', '_')
    s = re.sub(r'[^A-Za-z0-9א-ת_-]', '', s)
    return s or 'report'
